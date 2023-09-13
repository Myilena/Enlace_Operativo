import numpy as np
import openpyxl

# Genera 500 puntos de datos para la variable independiente X
X = np.random.rand(500, 1) * 10

# Define una ecuación lineal para la variable dependiente Y con un poco de ruido
a = 2.5
b = 10
error = np.random.randn(500, 1) * 2  # Agrega error aleatorio

Y = a * X + b + error

# Crear un nuevo archivo Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Datos de Regresión"

# Insertar encabezados
sheet.cell(row=1, column=1, value="Variable X")
sheet.cell(row=1, column=2, value="Variable Y")

# Insertar datos en Excel
for i in range(len(X)):
    sheet.cell(row=i + 2, column=1, value=X[i][0])
    sheet.cell(row=i + 2, column=2, value=Y[i][0])

# Guardar el archivo Excel
wb.save("datos_regresion.xlsx")
wb.close()

print("Datos insertados en el archivo Excel.")
